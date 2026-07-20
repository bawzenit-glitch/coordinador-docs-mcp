#!/usr/bin/env python3
"""
Servidor MCP — Documentos del Coordinador Eléctrico Nacional (multi-área).

Segundo conector, más amplio que `coordinador-mcp`: cubre TODAS las áreas de
"Informes y Estudios" del Coordinador (Planificación y Desarrollo, Operación,
Mercados, Reportes, Parámetros Operacionales, Auditorías Técnicas, Normativa
Sectorial) y además lee no solo PDF sino XLSX, CSV, TXT y ZIP.

El primer conector (Planificación y Desarrollo, solo PDF) sigue funcionando
intacto; este es independiente para poder compararlos.

Modelo del sitio
----------------
Cada página muestra "tarjetas":
  - CARPETA   -> título + enlace "Ver documentos"/"Ver informes" a otra página
  - DOCUMENTO -> título + (fecha opcional) + "Descargar PDF/XLSX/ZIP/..." a un archivo

Herramientas
------------
  listar_areas()                          -> las grandes áreas de documentos
  listar_secciones(area)                  -> categorías dentro de un área
  explorar(url)                           -> subcarpetas y documentos de una página
  arbol(url, profundidad)                 -> árbol anidado hasta cierta profundidad
  buscar_documentos(consulta, ...)        -> busca documentos por título en un árbol
  leer_documento(url, paginas)            -> lee PDF / XLSX / CSV / TXT / ZIP
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import tempfile
import time
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional
from urllib.parse import urljoin, urlparse

import httpx
from bs4 import BeautifulSoup
from mcp.server.fastmcp import FastMCP
from mcp.types import ToolAnnotations

# Herramientas de solo lectura: permite ejecucion automatica y desatendida.
RO = ToolAnnotations(readOnlyHint=True, openWorldHint=True)

# --------------------------------------------------------------------------- #
# Configuración
# --------------------------------------------------------------------------- #

BASE = "https://www.coordinador.cl"

USER_AGENT = "Mozilla/5.0 (compatible; CoordinadorDocsMCP/1.0; +https://www.coordinador.cl)"
REQUEST_TIMEOUT = 20.0
POLITE_DELAY = 0.05
# Presupuesto de tiempo por crawl: nunca corre más que esto, así una búsqueda
# en un área grande devuelve resultados parciales rápido y no deja el server
# "pegado" bloqueando otras llamadas.
CRAWL_TIME_BUDGET = float(os.environ.get("COORDINADOR_CRAWL_BUDGET", "20"))

CACHE_DIR = Path(
    os.environ.get("COORDINADOR_CACHE_DIR", Path(tempfile.gettempdir()) / "coordinador_docs_mcp")
)
CACHE_DIR.mkdir(parents=True, exist_ok=True)
INDEX_TTL = int(os.environ.get("COORDINADOR_INDEX_TTL", 24 * 3600))
CRAWL_MAX_PAGES = int(os.environ.get("COORDINADOR_MAX_PAGES", 600))

# Grandes áreas de "Informes y Estudios" (todas comparten el mismo patrón).
AREAS = [
    ("Planificación y Desarrollo", f"{BASE}/desarrollo/documentos/"),
    ("Operación", f"{BASE}/operacion/documentos/"),
    ("Mercados", f"{BASE}/mercados/documentos/"),
    ("Reportes del Coordinador", f"{BASE}/reportes/documentos/"),
    ("Parámetros Operacionales", f"{BASE}/parametros-operacionales/documentos/"),
    ("Auditorías Técnicas", f"{BASE}/auditorias-tecnicas/documentos/"),
    ("Normativa Sectorial", f"{BASE}/normativa-sectorial/documentos/"),
]

# Extensiones de archivo que consideramos "documentos".
DOC_EXTS = (
    ".pdf", ".xlsx", ".xlsm", ".xls", ".csv", ".txt",
    ".zip", ".doc", ".docx", ".ppt", ".pptx",
)

mcp = FastMCP(
    "coordinador-docs",
    host="0.0.0.0",
    port=int(os.environ.get("PORT", "8000")),
)

_client = httpx.Client(
    headers={"User-Agent": USER_AGENT},
    timeout=REQUEST_TIMEOUT,
    follow_redirects=True,
)

FECHA_RE = re.compile(r"Fecha de publicaci[oó]n:\s*([0-3]?\d/[0-1]?\d/\d{4})", re.I)

# --------------------------------------------------------------------------- #
# Utilidades
# --------------------------------------------------------------------------- #

def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def _fetch_html(url: str) -> str:
    r = _client.get(url)
    r.raise_for_status()
    return r.text


def _ext_of(url: str) -> str:
    return os.path.splitext(urlparse(url).path.lower())[1]


def _year_of(d: dict) -> Optional[str]:
    """Año del documento: de la fecha si existe; si no, del título o la URL.
    Muchos documentos (p. ej. de Operación) no traen fecha de publicación, pero
    su año aparece en el título ('EAF 234/2026') o en la ruta ('/2026/')."""
    fecha = d.get("fecha") or ""
    m = re.search(r"(\d{4})\s*$", fecha)
    if m:
        return m.group(1)
    for campo in (d.get("titulo", ""), d.get("url", "")):
        m = re.search(r"(20\d{2})", campo)
        if m:
            return m.group(1)
    return None


def _is_doc_link(text: str, url: str) -> bool:
    return text.startswith("descargar") or _ext_of(url) in DOC_EXTS


def _is_folder_link(text: str, url: str) -> bool:
    # Solo las tarjetas de contenido ("Ver documentos") son carpetas reales.
    # El footer "Ver Informes" enlaza a OTRAS áreas y se excluye a propósito
    # para no contaminar el crawl/índice.
    if text != "ver documentos":
        return False
    return "/documentos" in urlparse(url).path


def _area_prefix(url: str) -> str:
    """Prefijo del área, p. ej. '/operacion/documentos', para acotar el crawl."""
    parts = [p for p in urlparse(url).path.split("/") if p]
    return "/" + "/".join(parts[:2])


def _section_links(html: str, page_url: str, prefix: str) -> list[str]:
    """Enlaces del menú a secciones del área (un nivel bajo el prefijo)."""
    soup = BeautifulSoup(html, "html.parser")
    prefdepth = prefix.count("/")
    links, vistos = [], set()
    for a in soup.find_all("a", href=True):
        href = urljoin(page_url, a["href"]).split("#")[0]
        p = urlparse(href).path.rstrip("/")
        texto = _norm(a.get_text(" ")).lower()
        if (
            p.startswith(prefix + "/")
            and p.count("/") == prefdepth + 1
            and texto not in ("ver documentos", "ver informes")
            and href not in vistos
        ):
            vistos.add(href)
            links.append(href)
    return links


@dataclass
class Item:
    tipo: str            # "carpeta" | "documento"
    titulo: str
    url: str
    fecha: Optional[str] = None
    formato: Optional[str] = None


def _card_title(anchor) -> str:
    node = anchor
    for _ in range(5):
        node = node.parent
        if node is None:
            break
        h = node.find(["h1", "h2", "h3", "h4", "h5", "h6"])
        if h:
            t = _norm(h.get_text(" "))
            if t and "planificación y desarrollo" not in t.lower():
                return t
    node = anchor.parent
    txt = _norm(node.get_text(" ")) if node else ""
    txt = FECHA_RE.sub("", txt)
    for boiler in ("Ver documentos", "Ver informes", "Descargar PDF", "Descargar XLSX",
                   "Descargar ZIP", "Descargar CSV", "Descargar", "Fecha de publicación:"):
        txt = txt.replace(boiler, " ")
    return _norm(txt) or "(sin título)"


def _parse_items(html: str, page_url: str) -> list[Item]:
    soup = BeautifulSoup(html, "html.parser")
    items: list[Item] = []
    seen: set[tuple[str, str]] = set()

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        text = _norm(a.get_text(" ")).lower()
        abs_url = urljoin(page_url, href)

        if _is_doc_link(text, abs_url):
            key = ("documento", abs_url)
            if key in seen:
                continue
            seen.add(key)
            titulo = _card_title(a)
            fecha = None
            container = a.parent
            for _ in range(5):
                if container is None:
                    break
                m = FECHA_RE.search(container.get_text(" "))
                if m:
                    fecha = m.group(1)
                    break
                container = container.parent
            items.append(
                Item("documento", titulo, abs_url, fecha, (_ext_of(abs_url) or "").lstrip("."))
            )
            continue

        if _is_folder_link(text, abs_url):
            key = ("carpeta", abs_url)
            if key in seen:
                continue
            seen.add(key)
            items.append(Item("carpeta", _card_title(a), abs_url))

    return items


# --------------------------------------------------------------------------- #
# Índice de búsqueda
# --------------------------------------------------------------------------- #

def _index_path(root: str) -> Path:
    slug = re.sub(r"[^a-z0-9]+", "_", root.lower()).strip("_")
    return CACHE_DIR / f"index_{slug}.json"


def _crawl(root: str, max_pages: int = CRAWL_MAX_PAGES) -> tuple[list[dict], bool]:
    """Recorre el árbol del área. Devuelve (documentos, parcial). 'parcial' es
    True si el recorrido se detuvo por el presupuesto de tiempo o de páginas
    antes de vaciar la cola (quedó árbol sin visitar)."""
    docs: list[dict] = []
    visitados: set[str] = set()
    cola: list[tuple[str, list[str]]] = [(root, [])]
    prefix = _area_prefix(root)  # el crawl no sale de esta área
    n = 0
    seeded = False
    parcial = False
    t0 = time.monotonic()
    while cola and n < max_pages:
        if time.monotonic() - t0 > CRAWL_TIME_BUDGET:
            parcial = True
            break  # presupuesto agotado: devolvemos lo indexado hasta aquí
        url, ruta = cola.pop(0)
        url = url.split("#")[0].rstrip("/")
        if url in visitados:
            continue
        if not urlparse(url).path.startswith(prefix):
            continue  # nunca cruzar a otra área
        visitados.add(url)
        n += 1
        try:
            html = _fetch_html(url)
        except Exception:
            continue
        # Sembrar las secciones del menú una sola vez, para cubrir toda el área
        # aunque la raíz solo muestre "últimos documentos".
        if not seeded:
            seeded = True
            for sec in _section_links(html, url, prefix):
                cola.append((sec, ruta))
        for it in _parse_items(html, url):
            if it.tipo == "carpeta":
                cola.append((it.url, ruta + [it.titulo]))
            else:
                docs.append({
                    "titulo": it.titulo, "url": it.url, "fecha": it.fecha,
                    "formato": it.formato, "seccion": " / ".join(ruta) if ruta else "",
                    "pagina": url,
                })
        time.sleep(POLITE_DELAY)
    if cola:
        parcial = True  # quedó árbol sin visitar
    return docs, parcial


def _get_index(root: str, refrescar: bool = False) -> tuple[list[dict], bool]:
    """Devuelve (documentos, parcial), usando caché válida ~24 h."""
    path = _index_path(root)
    if not refrescar and path.exists():
        if time.time() - path.stat().st_mtime < INDEX_TTL:
            try:
                cache = json.loads(path.read_text(encoding="utf-8"))
                return cache.get("docs", []), cache.get("parcial", False)
            except Exception:
                pass
    docs, parcial = _crawl(root)
    path.write_text(
        json.dumps({"docs": docs, "parcial": parcial}, ensure_ascii=False),
        encoding="utf-8",
    )
    return docs, parcial


# --------------------------------------------------------------------------- #
# Lectores de archivos
# --------------------------------------------------------------------------- #

def _leer_pdf(content: bytes, paginas: Optional[str], max_chars: int) -> dict:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    n = len(reader.pages)
    indices = range(n)
    if paginas:
        m = re.match(r"\s*(\d+)\s*(?:-\s*(\d+))?\s*$", paginas)
        if m:
            a = int(m.group(1)); b = int(m.group(2)) if m.group(2) else a
            indices = range(max(0, a - 1), min(n, b))
    partes, total = [], 0
    for i in indices:
        try:
            t = reader.pages[i].extract_text() or ""
        except Exception:
            t = ""
        partes.append(f"\n--- Página {i + 1} ---\n{t}")
        total += len(t)
        if total >= max_chars:
            partes.append("\n[...truncado...]")
            break
    return {"tipo": "pdf", "n_paginas": n, "texto": "".join(partes)[:max_chars]}


def _leer_xlsx(content: bytes, max_chars: int, max_filas: int = 200) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    partes, total = [], 0
    for ws in wb.worksheets:
        partes.append(f"\n=== Hoja: {ws.title} ===")
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= max_filas:
                partes.append(f"[...más filas ({ws.max_row}) truncadas...]")
                break
            linea = "\t".join("" if c is None else str(c) for c in row)
            partes.append(linea)
            total += len(linea)
            if total >= max_chars:
                partes.append("\n[...truncado...]")
                wb.close()
                return {"tipo": "xlsx", "texto": "\n".join(partes)[:max_chars]}
    wb.close()
    return {"tipo": "xlsx", "hojas": len(wb.sheetnames) if hasattr(wb, "sheetnames") else None,
            "texto": "\n".join(partes)[:max_chars]}


def _leer_csv_txt(content: bytes, max_chars: int) -> dict:
    for enc in ("utf-8", "latin-1"):
        try:
            texto = content.decode(enc)
            break
        except Exception:
            texto = content.decode("utf-8", errors="replace")
    return {"tipo": "texto", "texto": texto[:max_chars]}


def _leer_zip(content: bytes, max_chars: int, archivo: Optional[str] = None) -> dict:
    """Sin 'archivo': lista lo que trae el ZIP (nombre, tamaño, formato) sin
    volcar contenido. Con 'archivo': extrae y lee ese archivo interno."""
    zf = zipfile.ZipFile(io.BytesIO(content))
    infos = zf.infolist()

    if archivo is None:
        return {
            "tipo": "zip",
            "n_archivos": len(infos),
            "archivos": [
                {"nombre": i.filename, "bytes": i.file_size,
                 "formato": os.path.splitext(i.filename.lower())[1].lstrip(".")}
                for i in infos if not i.is_dir()
            ],
            "nota": "Para leer uno, llama leer_documento con archivo_zip=<nombre>.",
        }

    # Buscar el archivo interno (coincidencia exacta o por sufijo/nombre)
    candidatos = [i.filename for i in infos if not i.is_dir()]
    elegido = next(
        (n for n in candidatos if n == archivo or n.endswith("/" + archivo)
         or os.path.basename(n) == archivo),
        None,
    )
    if elegido is None:
        return {"tipo": "zip", "error": f"'{archivo}' no está en el ZIP.",
                "archivos": [i.filename for i in infos if not i.is_dir()]}

    data = zf.read(elegido)
    ext = os.path.splitext(elegido.lower())[1]
    if ext == ".pdf":
        sub = _leer_pdf(data, None, max_chars)
    elif ext in (".xlsx", ".xlsm"):
        sub = _leer_xlsx(data, max_chars)
    elif ext in (".csv", ".txt"):
        sub = _leer_csv_txt(data, max_chars)
    else:
        sub = {"tipo": ext.lstrip("."), "nota": "Formato no legible como texto.",
               "bytes": len(data)}
    sub["archivo_zip"] = elegido
    return sub


# --------------------------------------------------------------------------- #
# Herramientas MCP
# --------------------------------------------------------------------------- #

@mcp.tool(annotations=RO)
def listar_areas() -> str:
    """
    Lista las grandes áreas de documentos del Coordinador (Planificación y
    Desarrollo, Operación, Mercados, Reportes, Parámetros Operacionales,
    Auditorías Técnicas, Normativa Sectorial), con su URL. Punto de partida.
    """
    return json.dumps(
        {"areas": [{"nombre": n, "url": u} for n, u in AREAS], "total": len(AREAS)},
        ensure_ascii=False, indent=2,
    )


@mcp.tool(annotations=RO)
def listar_secciones(area: str) -> str:
    """
    Lista las categorías (secciones) dentro de un área, leyendo el menú del
    Coordinador. Usa una URL de listar_areas.

    Args:
        area: URL del área (p. ej. https://www.coordinador.cl/operacion/documentos/).
    """
    area = area.strip()
    if not area.startswith("http"):
        area = urljoin(BASE, area)
    base_path = urlparse(area).path.rstrip("/")  # p. ej. /operacion/documentos
    secciones, vistos = [], set()
    try:
        soup = BeautifulSoup(_fetch_html(area), "html.parser")
        for a in soup.find_all("a", href=True):
            href = urljoin(area, a["href"])
            texto = _norm(a.get_text(" "))
            p = urlparse(href).path.rstrip("/")
            if (
                p.startswith(base_path + "/")
                and p.count("/") == base_path.count("/") + 1
                and texto
                and texto.lower() not in ("ver documentos", "ver informes")
                and href not in vistos
            ):
                vistos.add(href)
                secciones.append({"nombre": texto, "url": href})
    except Exception as e:
        return json.dumps({"error": f"No se pudo cargar el área: {e}", "url": area}, ensure_ascii=False)
    return json.dumps({"area": area, "secciones": secciones, "total": len(secciones)},
                      ensure_ascii=False, indent=2)


@mcp.tool(annotations=RO)
def explorar(url: str) -> str:
    """
    Explora UNA página del repositorio y devuelve sus tarjetas: subcarpetas y
    documentos (título, fecha, formato y enlace directo al archivo). Los
    documentos pueden ser PDF, XLSX, ZIP, CSV, etc.

    Args:
        url: URL de una página bajo /.../documentos/ (área, sección o subcarpeta).
    """
    url = url.strip()
    if not url.startswith("http"):
        url = urljoin(BASE, url)
    try:
        html = _fetch_html(url)
    except Exception as e:
        return json.dumps({"error": f"No se pudo cargar: {e}", "url": url}, ensure_ascii=False)
    items = _parse_items(html, url)
    carpetas = [{"titulo": i.titulo, "url": i.url} for i in items if i.tipo == "carpeta"]
    documentos = [
        {"titulo": i.titulo, "url": i.url, "fecha": i.fecha, "formato": i.formato}
        for i in items if i.tipo == "documento"
    ]
    return json.dumps(
        {"url": url, "subcarpetas": carpetas, "documentos": documentos,
         "n_subcarpetas": len(carpetas), "n_documentos": len(documentos)},
        ensure_ascii=False, indent=2,
    )


@mcp.tool(annotations=RO)
def arbol(url: str, profundidad: int = 2) -> str:
    """
    Devuelve el árbol anidado de carpetas y documentos desde una página, hasta
    la profundidad indicada (1-4). Útil para ver la estructura sin leer archivos.
    """
    url = url.strip()
    if not url.startswith("http"):
        url = urljoin(BASE, url)
    profundidad = max(1, min(int(profundidad), 4))
    visitados: set[str] = set()

    def expandir(u: str, prof: int) -> dict:
        u = u.split("#")[0].rstrip("/")
        nodo = {"url": u, "documentos": [], "subcarpetas": []}
        if u in visitados:
            return nodo
        visitados.add(u)
        try:
            items = _parse_items(_fetch_html(u), u)
        except Exception:
            return nodo
        for it in items:
            if it.tipo == "documento":
                nodo["documentos"].append({"titulo": it.titulo, "url": it.url,
                                           "fecha": it.fecha, "formato": it.formato})
            elif prof > 1:
                sub = expandir(it.url, prof - 1)
                sub["titulo"] = it.titulo
                nodo["subcarpetas"].append(sub)
            else:
                nodo["subcarpetas"].append({"titulo": it.titulo, "url": it.url})
        time.sleep(POLITE_DELAY)
        return nodo

    return json.dumps(expandir(url, profundidad), ensure_ascii=False, indent=2)


@mcp.tool(annotations=RO)
def buscar_documentos(
    consulta: str,
    seccion: str,
    anio: Optional[int] = None,
    formato: Optional[str] = None,
    limite: int = 30,
    refrescar_indice: bool = False,
) -> str:
    """
    Busca documentos por título en todo el árbol de una sección o área. La
    primera llamada construye un índice recorriendo esa rama (puede tardar en
    áreas grandes como Operación); luego usa caché (~24 h).

    Args:
        consulta: términos a buscar (p. ej. "informe final", "EAF", "costo marginal").
        seccion: URL del área o sección donde buscar (de listar_areas / listar_secciones).
        anio: filtra por año (de la fecha, o del título/URL si no hay fecha).
        formato: filtra por extensión (p. ej. "pdf", "xlsx", "zip").
        limite: máximo de resultados (por defecto 30).
        refrescar_indice: fuerza reconstruir el índice.

    Nota: en áreas grandes la primera pasada indexa dentro de un presupuesto de
    tiempo; si el resultado trae "indice_parcial": true, acota con una URL de
    listar_secciones o usa refrescar_indice para profundizar.
    """
    root = seccion.strip()
    if not root.startswith("http"):
        root = urljoin(BASE, root)
    try:
        docs, parcial = _get_index(root, refrescar=refrescar_indice)
    except Exception as e:
        return json.dumps({"error": f"No se pudo construir el índice: {e}"}, ensure_ascii=False)

    terminos = [t for t in re.split(r"\s+", consulta.lower().strip()) if t]

    def coincide(d: dict) -> bool:
        blob = f"{d.get('titulo','')} {d.get('seccion','')}".lower()
        if not all(t in blob for t in terminos):
            return False
        if anio is not None and _year_of(d) != str(anio):
            return False
        if formato is not None and (d.get("formato") or "").lower() != formato.lower().lstrip("."):
            return False
        return True

    resultados = [d for d in docs if coincide(d)]

    def clave(d):
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", d.get("fecha") or "")
        if m:
            return (m.group(3), m.group(2), m.group(1))
        return (_year_of(d) or "0000", "00", "00")  # sin fecha: ordena por año inferido

    resultados.sort(key=clave, reverse=True)
    return json.dumps(
        {"consulta": consulta, "seccion": root, "total_indexado": len(docs),
         "indice_parcial": parcial, "coincidencias": len(resultados),
         "resultados": resultados[: max(1, int(limite))]},
        ensure_ascii=False, indent=2,
    )


@mcp.tool(annotations=RO)
def leer_documento(
    url: str,
    paginas: Optional[str] = None,
    archivo_zip: Optional[str] = None,
    max_chars: int = 60000,
) -> str:
    """
    Descarga y lee un archivo del Coordinador. Soporta PDF, XLSX/XLSM, CSV, TXT
    y ZIP. Para un ZIP, por defecto LISTA los archivos que contiene; para leer
    uno, pásalo en 'archivo_zip'. Para otros formatos devuelve solo metadatos.

    Args:
        url: URL directa del archivo (campo "url" de un documento).
        paginas: solo PDF; rango como "1-5" o "3".
        archivo_zip: nombre de un archivo dentro del ZIP para extraer y leer.
        max_chars: tope de caracteres devueltos.
    """
    url = url.strip()
    ext = _ext_of(url)
    try:
        r = _client.get(url)
        r.raise_for_status()
        content = r.content
    except Exception as e:
        return json.dumps({"error": f"No se pudo descargar: {e}", "url": url}, ensure_ascii=False)

    try:
        if ext == ".pdf":
            out = _leer_pdf(content, paginas, max_chars)
        elif ext in (".xlsx", ".xlsm"):
            out = _leer_xlsx(content, max_chars)
        elif ext in (".csv", ".txt"):
            out = _leer_csv_txt(content, max_chars)
        elif ext == ".zip":
            out = _leer_zip(content, max_chars, archivo=archivo_zip)
        else:
            out = {"tipo": ext.lstrip("."), "nota": "Formato no legible como texto; solo descarga.",
                   "bytes": len(content)}
    except Exception as e:
        return json.dumps({"error": f"No se pudo leer el archivo ({ext}): {e}", "url": url}, ensure_ascii=False)

    out["url"] = url
    return json.dumps(out, ensure_ascii=False)


if __name__ == "__main__":
    transport = os.environ.get("MCP_TRANSPORT", "streamable-http")
    if transport == "stdio":
        mcp.run()
    else:
        mcp.run(transport="streamable-http")
